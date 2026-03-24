import csv
import io
import os
import uuid
from calendar import monthrange
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (Blueprint, render_template, redirect, url_for, flash,
                   request, make_response, send_from_directory, current_app)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from ..models import db, Booking, Invoice, Expense, BankTransaction, EXPENSE_CATEGORIES
from ..utils import hotel_date

accounting_bp = Blueprint('accounting', __name__, url_prefix='/accounting')

ALLOWED_RECEIPT = {'jpg', 'jpeg', 'png', 'pdf'}
TOTAL_ROOMS = 8


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Admin access required.', 'error')
            return redirect(url_for('bookings.index'))
        return f(*args, **kwargs)
    return decorated


def _allowed_receipt(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_RECEIPT


def _save_receipt(file):
    ext = file.filename.rsplit('.', 1)[1].lower()
    name = f'receipt_{uuid.uuid4().hex[:10]}.{ext}'
    upload_dir = os.path.join(current_app.root_path, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    file.save(os.path.join(upload_dir, name))
    return name


def _month_revenue(year, month):
    """Total amount_paid on invoices whose booking check_in_date is in given month."""
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    return db.session.query(db.func.coalesce(db.func.sum(Invoice.amount_paid), 0)).join(Booking).filter(
        Invoice.payment_status.in_(['paid', 'partial']),
        Booking.check_in_date >= start,
        Booking.check_in_date <= end,
    ).scalar()


def _month_occupancy(year, month):
    """Occupancy percentage for a given month."""
    start = date(year, month, 1)
    days_in_month = monthrange(year, month)[1]
    end = date(year, month, days_in_month)

    bookings = Booking.query.filter(
        Booking.status.in_(['checked_in', 'checked_out']),
        Booking.check_in_date <= end,
        Booking.check_out_date > start,
    ).all()

    total_nights = 0
    for b in bookings:
        effective_in = max(b.check_in_date, start)
        effective_out = min(b.check_out_date, end + timedelta(days=1))
        nights = (effective_out - effective_in).days
        if nights > 0:
            total_nights += nights

    capacity = TOTAL_ROOMS * days_in_month
    return round(total_nights / capacity * 100, 1) if capacity else 0


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@accounting_bp.route('/')
@login_required
@admin_required
def dashboard():
    today = hotel_date()

    # Last 12 months bar chart data
    months_labels = []
    months_revenue = []
    for i in range(11, -1, -1):
        d = date(today.year, today.month, 1) - timedelta(days=i * 28)
        # normalise to first of month
        d = date(d.year, d.month, 1)
        label = d.strftime('%b %Y')
        rev = float(_month_revenue(d.year, d.month))
        months_labels.append(label)
        months_revenue.append(rev)

    # Deduplicate (timedelta trick may repeat months near boundaries)
    seen = {}
    for lbl, rev in zip(months_labels, months_revenue):
        seen[lbl] = rev
    # Rebuild last-12 unique
    months_labels = list(seen.keys())[-12:]
    months_revenue = list(seen.values())[-12:]

    # Current month & YTD revenue
    current_month_revenue = float(_month_revenue(today.year, today.month))
    ytd_revenue = sum(
        float(_month_revenue(today.year, m)) for m in range(1, today.month + 1)
    )

    # Occupancy last 6 months
    occupancy_data = []
    for i in range(5, -1, -1):
        d = date(today.year, today.month, 1) - timedelta(days=i * 28)
        d = date(d.year, d.month, 1)
        occupancy_data.append({
            'label': d.strftime('%b %Y'),
            'rate': _month_occupancy(d.year, d.month),
        })

    # Revenue forecast next 3 months
    forecast = []
    for i in range(1, 4):
        fd = date(today.year, today.month, 1)
        # Add i months
        month = today.month + i
        year = today.year + (month - 1) // 12
        month = (month - 1) % 12 + 1
        fd = date(year, month, 1)
        end = date(year, month, monthrange(year, month)[1])
        total = db.session.query(
            db.func.coalesce(db.func.sum(Booking.total_amount), 0)
        ).filter(
            Booking.status.in_(['confirmed', 'checked_in']),
            Booking.check_in_date >= fd,
            Booking.check_in_date <= end,
        ).scalar()
        forecast.append({'label': fd.strftime('%B %Y'), 'amount': float(total)})

    # Outstanding invoices
    outstanding = Invoice.query.filter(
        Invoice.payment_status.in_(['unpaid', 'partial'])
    ).join(Booking).order_by(Invoice.created_at.desc()).all()
    outstanding_total = sum(i.balance_due for i in outstanding)

    return render_template(
        'accounting/dashboard.html',
        months_labels=months_labels,
        months_revenue=months_revenue,
        current_month_revenue=current_month_revenue,
        ytd_revenue=ytd_revenue,
        occupancy_data=occupancy_data,
        forecast=forecast,
        outstanding=outstanding,
        outstanding_total=outstanding_total,
    )


# ─── EXPENSES ────────────────────────────────────────────────────────────────

@accounting_bp.route('/expenses/')
@login_required
@admin_required
def expenses():
    cat_filter = request.args.get('category', '')
    query = Expense.query
    if cat_filter:
        query = query.filter(Expense.category == cat_filter)
    expense_list = query.order_by(Expense.date.desc()).all()
    return render_template(
        'accounting/expenses.html',
        expenses=expense_list,
        categories=EXPENSE_CATEGORIES,
        cat_filter=cat_filter,
    )


@accounting_bp.route('/expenses/add', methods=['POST'])
@login_required
@admin_required
def add_expense():
    try:
        exp_date = date.fromisoformat(request.form.get('date', ''))
        amount = float(request.form.get('amount', 0))
    except (ValueError, TypeError):
        flash('Invalid date or amount.', 'error')
        return redirect(url_for('accounting.expenses'))

    category = request.form.get('category', '').strip()
    if category not in EXPENSE_CATEGORIES:
        flash('Invalid category.', 'error')
        return redirect(url_for('accounting.expenses'))

    receipt_filename = None
    receipt_file = request.files.get('receipt')
    if receipt_file and receipt_file.filename and _allowed_receipt(receipt_file.filename):
        receipt_filename = _save_receipt(receipt_file)

    expense = Expense(
        date=exp_date,
        category=category,
        amount=amount,
        description=request.form.get('description', '').strip() or None,
        receipt_filename=receipt_filename,
        created_by=current_user.id,
    )
    db.session.add(expense)
    db.session.commit()
    flash(f'Expense of MVR {amount:.0f} ({category}) recorded.', 'success')
    return redirect(url_for('accounting.expenses'))


@accounting_bp.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    db.session.delete(expense)
    db.session.commit()
    flash('Expense deleted.', 'success')
    return redirect(url_for('accounting.expenses'))


@accounting_bp.route('/receipts/<path:filename>')
@login_required
@admin_required
def download_receipt(filename):
    upload_dir = os.path.join(current_app.root_path, 'uploads')
    full_path = os.path.join(upload_dir, filename)
    if not os.path.isfile(full_path):
        flash('Receipt file not found.', 'error')
        return redirect(url_for('accounting.expenses'))
    return send_from_directory(upload_dir, filename)


# ─── BANK RECONCILIATION ─────────────────────────────────────────────────────

def _try_parse_date(s):
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


@accounting_bp.route('/reconciliation/')
@login_required
@admin_required
def reconciliation():
    transactions = BankTransaction.query.order_by(BankTransaction.statement_date.desc()).all()
    matched = sum(1 for t in transactions if t.match_type != 'unmatched')
    unmatched = sum(1 for t in transactions if t.match_type == 'unmatched')
    return render_template(
        'accounting/reconciliation.html',
        transactions=transactions,
        matched=matched,
        unmatched=unmatched,
    )


@accounting_bp.route('/reconciliation/upload', methods=['POST'])
@login_required
@admin_required
def upload_reconciliation():
    csv_file = request.files.get('statement')
    if not csv_file or not csv_file.filename:
        flash('Please select a CSV file.', 'error')
        return redirect(url_for('accounting.reconciliation'))
    if not csv_file.filename.lower().endswith('.csv'):
        flash('Only CSV files are supported. Export your bank statement as CSV.', 'error')
        return redirect(url_for('accounting.reconciliation'))

    content = csv_file.read().decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(content))

    # Delete existing transactions
    BankTransaction.query.delete()

    # Normalise column names to lowercase
    rows = []
    errors = 0
    for row in reader:
        row_lower = {k.lower().strip(): v for k, v in row.items()}
        # Find date column
        date_val = None
        for col in ('date', 'transaction date', 'value date', 'posting date'):
            if col in row_lower and row_lower[col].strip():
                date_val = _try_parse_date(row_lower[col])
                if date_val:
                    break
        # Find amount column
        amount_val = None
        for col in ('amount', 'credit', 'debit', 'value'):
            if col in row_lower and row_lower[col].strip():
                try:
                    raw = row_lower[col].replace(',', '').replace(' ', '')
                    amount_val = float(raw)
                    break
                except ValueError:
                    pass
        # Find description
        desc_val = ''
        for col in ('description', 'details', 'narrative', 'particulars', 'reference'):
            if col in row_lower and row_lower[col].strip():
                desc_val = row_lower[col].strip()
                break

        if not date_val or amount_val is None:
            errors += 1
            continue
        rows.append((date_val, desc_val, amount_val))

    # Match each row
    for stmt_date, desc, amount in rows:
        match_type = 'unmatched'
        match_ref = None

        # Try invoice match: same amount_paid and invoice issue_date within ±1 day
        for delta in (0, 1, -1):
            check_date = stmt_date + timedelta(days=delta)
            inv = Invoice.query.filter(
                Invoice.issue_date == check_date,
                db.func.abs(Invoice.amount_paid - amount) < 0.01,
                Invoice.payment_status.in_(['paid', 'partial']),
            ).first()
            if inv:
                match_type = 'invoice'
                match_ref = inv.invoice_number
                break

        if match_type == 'unmatched':
            for delta in (0, 1, -1):
                check_date = stmt_date + timedelta(days=delta)
                exp = Expense.query.filter(
                    Expense.date == check_date,
                    db.func.abs(Expense.amount - abs(amount)) < 0.01,
                ).first()
                if exp:
                    match_type = 'expense'
                    match_ref = exp.category
                    break

        db.session.add(BankTransaction(
            statement_date=stmt_date,
            description=desc,
            amount=amount,
            match_type=match_type,
            match_ref=match_ref,
        ))

    db.session.commit()
    total = len(rows)
    matched = sum(1 for _ in BankTransaction.query.filter(BankTransaction.match_type != 'unmatched').all())
    flash(
        f'Parsed {total} transactions ({matched} matched, {total - matched} unmatched). '
        f'{errors} rows skipped due to missing date/amount.',
        'success' if errors == 0 else 'warning',
    )
    return redirect(url_for('accounting.reconciliation'))


# ─── PROFIT & LOSS ───────────────────────────────────────────────────────────

@accounting_bp.route('/pl/')
@login_required
@admin_required
def pl():
    today = hotel_date()
    try:
        year = int(request.args.get('year', today.year))
        month = int(request.args.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])

    # Revenue
    revenue = float(db.session.query(
        db.func.coalesce(db.func.sum(Invoice.amount_paid), 0)
    ).join(Booking).filter(
        Invoice.payment_status.in_(['paid', 'partial']),
        Booking.check_in_date >= start,
        Booking.check_in_date <= end,
    ).scalar())

    # Expenses by category
    expense_rows = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount).label('total')
    ).filter(
        Expense.date >= start,
        Expense.date <= end,
    ).group_by(Expense.category).all()

    expenses_by_cat = {row.category: float(row.total) for row in expense_rows}
    total_expenses = sum(expenses_by_cat.values())
    net_profit = revenue - total_expenses

    # Prev/next month links
    prev_month = month - 1 or 12
    prev_year = year if month > 1 else year - 1
    next_month = month % 12 + 1
    next_year = year if month < 12 else year + 1

    return render_template(
        'accounting/pl.html',
        year=year, month=month,
        month_label=start.strftime('%B %Y'),
        revenue=revenue,
        expenses_by_cat=expenses_by_cat,
        total_expenses=total_expenses,
        net_profit=net_profit,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        categories=EXPENSE_CATEGORIES,
    )


# ─── TAX SUMMARY ─────────────────────────────────────────────────────────────

TAX_RATE = 0.0  # Update when tax applies


@accounting_bp.route('/tax/')
@login_required
@admin_required
def tax():
    today = hotel_date()
    try:
        year = int(request.args.get('year', today.year))
        month = int(request.args.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])

    total_collected = float(db.session.query(
        db.func.coalesce(db.func.sum(Invoice.amount_paid), 0)
    ).join(Booking).filter(
        Invoice.payment_status.in_(['paid', 'partial']),
        Booking.check_in_date >= start,
        Booking.check_in_date <= end,
    ).scalar())

    tax_payable = round(total_collected * TAX_RATE, 2)

    prev_month = month - 1 or 12
    prev_year = year if month > 1 else year - 1
    next_month = month % 12 + 1
    next_year = year if month < 12 else year + 1

    return render_template(
        'accounting/tax.html',
        year=year, month=month,
        month_label=start.strftime('%B %Y'),
        total_collected=total_collected,
        tax_rate=TAX_RATE,
        tax_payable=tax_payable,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
    )


# ─── REPORTS ─────────────────────────────────────────────────────────────────

@accounting_bp.route('/reports/')
@login_required
@admin_required
def reports():
    today = hotel_date()
    try:
        year = int(request.args.get('year', today.year))
        month = int(request.args.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    start = date(year, month, 1)
    return render_template(
        'accounting/reports.html',
        year=year, month=month,
        month_label=start.strftime('%B %Y'),
    )


@accounting_bp.route('/reports/pdf')
@login_required
@admin_required
def export_pdf():
    from ..services.accounting_pdf import generate_monthly_report_pdf
    today = hotel_date()
    try:
        year = int(request.args.get('year', today.year))
        month = int(request.args.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        year, month = today.year, today.month

    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])

    revenue = float(db.session.query(
        db.func.coalesce(db.func.sum(Invoice.amount_paid), 0)
    ).join(Booking).filter(
        Invoice.payment_status.in_(['paid', 'partial']),
        Booking.check_in_date >= start,
        Booking.check_in_date <= end,
    ).scalar())

    expense_rows = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount).label('total')
    ).filter(
        Expense.date >= start,
        Expense.date <= end,
    ).group_by(Expense.category).all()

    expenses_by_cat = {row.category: float(row.total) for row in expense_rows}
    total_expenses = sum(expenses_by_cat.values())
    net_profit = revenue - total_expenses

    buf = generate_monthly_report_pdf(
        year=year, month=month,
        month_label=start.strftime('%B %Y'),
        revenue=revenue,
        expenses_by_cat=expenses_by_cat,
        total_expenses=total_expenses,
        net_profit=net_profit,
        generated_on=today,
    )
    filename = f'sheeza-manzil-pl-{year}-{month:02d}.pdf'
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@accounting_bp.route('/reports/excel')
@login_required
@admin_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    headers = ['Date', 'Category', 'Amount (MVR)', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    expenses = Expense.query.order_by(Expense.date.desc()).all()
    for row_idx, exp in enumerate(expenses, 2):
        ws.cell(row=row_idx, column=1, value=exp.date)
        ws.cell(row=row_idx, column=1).number_format = 'YYYY-MM-DD'
        ws.cell(row=row_idx, column=2, value=exp.category)
        ws.cell(row=row_idx, column=3, value=exp.amount)
        ws.cell(row=row_idx, column=4, value=exp.description or '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = hotel_date()
    filename = f'sheeza-manzil-expenses-{today}.xlsx'
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
